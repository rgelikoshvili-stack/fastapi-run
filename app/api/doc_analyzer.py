from __future__ import annotations
from dataclasses import dataclass, asdict
from typing import Any, Dict, List
import re, io, json, base64, urllib.request

try:
    import fitz
except Exception:
    fitz = None
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

def _pdf_to_text(data):
    if fitz is None: return ""
    doc = fitz.open(stream=data, filetype="pdf")
    return "\n".join(doc[i].get_text("text") or "" for i in range(len(doc))).strip()

def _ocr_with_vision(data):
    try:
        req = urllib.request.Request(
            "http://metadata.google.internal/computeMetadata/v1/instance/service-accounts/default/token",
            headers={"Metadata-Flavor": "Google"}
        )
        with urllib.request.urlopen(req, timeout=5) as r:
            token = json.loads(r.read())["access_token"]
        if fitz is None: return "OCR_NO_FITZ"
        doc = fitz.open(stream=data, filetype="pdf")
        texts = []
        for i in range(len(doc)):
            pix = doc[i].get_pixmap(matrix=fitz.Matrix(2, 2))
            img_b64 = base64.b64encode(pix.tobytes("png")).decode()
            body = json.dumps({"requests":[{"image":{"content":img_b64},"features":[{"type":"DOCUMENT_TEXT_DETECTION"}]}]}).encode()
            req2 = urllib.request.Request(
                "https://vision.googleapis.com/v1/images:annotate",
                data=body,
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
            )
            with urllib.request.urlopen(req2, timeout=30) as r2:
                resp = json.loads(r2.read())
            t = resp.get("responses",[{}])[0].get("fullTextAnnotation",{}).get("text","")
            texts.append(t or f"OCR_NO_TEXT_PAGE_{i}")
        return "\n".join(texts) if texts else "OCR_EMPTY"
    except Exception as e:
        return f"OCR_ERROR: {e}"

def _xlsx_to_text(data):
    if load_workbook is None: return ""
    wb = load_workbook(io.BytesIO(data), data_only=True)
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
            if vals: out.append(" | ".join(vals))
    return "\n".join(out).strip()

def _raw_to_text(data):
    return data.decode("utf-8", errors="ignore").strip()

def extract_text(filename, data):
    n = (filename or "").lower()
    if n.endswith(".pdf") or data[:4] == b"%PDF":
        text = _pdf_to_text(data)
        if not text or len(text) < 50:
            text = _ocr_with_vision(data)
        return "pdf", text
    if n.endswith((".xlsx", ".xlsm")): return "xlsx", _xlsx_to_text(data)
    return "text", _raw_to_text(data)

DATE_RE = [r"\b(\d{2})[./-](\d{2})[./-](\d{4})\b", r"\b(\d{4})[./-](\d{2})[./-](\d{2})\b", r"\b(\d{2}),(\d{2}),(\d{4})\b"]
PERIOD_RE = [r"\b(\d{2}[./-]\d{2}[./-]\d{4})\s*[-–]\s*(\d{2}[./-]\d{2}[./-]\d{4})\b"]
AMOUNT_RE = r"(?<!\w)(\d{1,3}(?:[ ,]\d{3})*(?:[.,]\d{2})?|\d+(?:[.,]\d{2})?)(?:\s*(GEL|USD|EUR|₾|\$|€))?(?!\w)"
ID_RE = r"\b(\d{9}|\d{11})\b"
IBAN_RE = r"\bGE\d{2}[A-Z]{2}\d{16}\b"
NOISE = {"2020","2021","2022","2023","2024","2025","2026","2027","2028"}
TERMS = ["ხელშეკრულება","ვადა","გადახდა","ჯარიმა","პირგასამტეხლო","Payment","Due","Penalty","срок","штраф"]

def extract_dates(text):
    out = set()
    for pat in DATE_RE:
        for m in re.finditer(pat, text): out.add(m.group(0))
    return sorted(out)

def extract_periods(text):
    out = set()
    for pat in PERIOD_RE:
        for m in re.finditer(pat, text): out.add(m.group(0))
    return sorted(out)

def extract_amounts(text):
    out, seen = [], set()
    for m in re.finditer(AMOUNT_RE, text):
        s = m.group(1).strip().replace(" ","").replace(",","")
        try: val = float(s)
        except: continue
        if val == 0: continue
        cur = m.group(2)
        if cur == "₾": cur = "GEL"
        if cur == "$": cur = "USD"
        if cur == "€": cur = "EUR"
        k = (val, cur, m.group(0))
        if k in seen: continue
        seen.add(k)
        out.append({"raw": m.group(0), "value": val, "currency": cur})
    return out

def extract_ids(text):
    return sorted(set(re.findall(ID_RE, text)) - NOISE)

def extract_ibans(text):
    return sorted(set(re.findall(IBAN_RE, text)))

def extract_names(text):
    names = set()
    for ln in text.splitlines()[:4000]:
        ln = ln.strip()
        if any(k in ln for k in ["შპს","ი.მ","ი/მ","LLC","Ltd","ООО"]):
            names.add(ln[:160])
    return sorted(names)[:50]

def extract_terms(text):
    result = {"has_terms": False, "snippets": [], "deadlines": []}
    if any(k.lower() in text.lower() for k in TERMS):
        result["has_terms"] = True
    for ch in re.split(r"(?<=[\.\!\?])\s+|\n+", text):
        c = ch.strip()
        if c and any(k.lower() in c.lower() for k in TERMS):
            result["snippets"].append(c[:400])
        if len(result["snippets"]) >= 20: break
    return result

@dataclass
class AnalysisResult:
    filename: str
    doc_format: str
    text_len: int
    ocr_used: bool
    dates: List[str]
    periods: List[str]
    amounts: List[Dict[str, Any]]
    ids: List[str]
    ibans: List[str]
    names: List[str]
    terms: Dict[str, Any]
    warnings: List[str]

def analyze(filename, data):
    doc_format, text = extract_text(filename, data)
    ocr_used = False
    warnings = []
    n = (filename or "").lower()
    if n.endswith(".pdf") or data[:4] == b"%PDF":
        native = _pdf_to_text(data)
        if not native or len(native) < 50:
            ocr_used = True
    if not text or text.startswith("OCR_"): warnings.append(f"No text extracted: {text[:100]}")
    if not extract_amounts(text): warnings.append("No amounts detected")
    if not extract_dates(text): warnings.append("No dates detected")
    return AnalysisResult(
        filename=filename, doc_format=doc_format, text_len=len(text),
        ocr_used=ocr_used,
        dates=extract_dates(text), periods=extract_periods(text),
        amounts=extract_amounts(text), ids=extract_ids(text),
        ibans=extract_ibans(text), names=extract_names(text),
        terms=extract_terms(text), warnings=warnings
    )

def to_dict(res):
    return asdict(res)