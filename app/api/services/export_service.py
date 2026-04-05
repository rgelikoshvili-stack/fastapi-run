import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import psycopg2.extras
from app.api.db import get_db


def get_journal_drafts(tenant_id="default", status=None, limit=1000):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        if status:
            cur.execute("""
                SELECT id, date, description, partner, amount,
                       debit_account, credit_account, account_code,
                       reason, confidence, status, source_type, created_at
                FROM journal_drafts
                WHERE tenant_id = %s AND status = %s
                ORDER BY created_at DESC LIMIT %s
            """, (tenant_id, status, limit))
        else:
            cur.execute("""
                SELECT id, date, description, partner, amount,
                       debit_account, credit_account, account_code,
                       reason, confidence, status, source_type, created_at
                FROM journal_drafts
                WHERE tenant_id = %s
                ORDER BY created_at DESC LIMIT %s
            """, (tenant_id, limit))
        return [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()


def export_excel(tenant_id="default", status=None):
    rows = get_journal_drafts(tenant_id, status)
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Drafts"

    ws["A1"] = "Bridge Hub — Journal Drafts"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Tenant: {tenant_id} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10)

    headers = ["ID", "თარიღი", "აღწერა", "პარტნიორი", "თანხა",
               "დებეტი", "კრედიტი", "ანგ. კოდი", "მიზეზი",
               "Confidence", "სტატუსი", "წყარო", "შექმნილი"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center")

    status_colors = {
        "approved": "E2EFDA",
        "auto_approved": "DDEBF7",
        "rejected": "FCE4D6",
        "pending_approval": "FFF2CC",
    }

    for row_idx, r in enumerate(rows, 5):
        fill_color = status_colors.get(r.get("status", ""), "FFFFFF")
        values = [
            r.get("id"), str(r.get("date", ""))[:10],
            r.get("description", ""), r.get("partner", ""),
            r.get("amount"), r.get("debit_account"),
            r.get("credit_account"), r.get("account_code"),
            r.get("reason"), r.get("confidence"),
            r.get("status"), r.get("source_type"),
            str(r.get("created_at", ""))[:19]
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=fill_color)

    for letter, w in zip(['A','B','C','D','E','F','G','H','I','J','K','L','M'],
                         [8, 12, 35, 20, 12, 10, 10, 12, 15, 12, 15, 10, 18]):
        ws.column_dimensions[letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_csv_drafts(tenant_id="default", status=None):
    rows = get_journal_drafts(tenant_id, status)
    output = io.StringIO()
    fields = ["id", "date", "description", "partner", "amount",
              "debit_account", "credit_account", "account_code",
              "reason", "confidence", "status", "source_type", "created_at"]
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow({k: r.get(k, "") for k in fields})
    output.seek(0)
    return output


def export_1c_xml(tenant_id="default", status="approved"):
    rows = get_journal_drafts(tenant_id, status)
    lines = ['<?xml version="1.0" encoding="UTF-8"?>']
    lines.append(f'<ФайлОбменДанными ВерсияФормата="1.0" ДатаФормирования="{datetime.now().strftime("%Y-%m-%d")}">')
    lines.append('  <Документы>')
    for r in rows:
        lines.append('  <ДокументБухгалтерскийУчет>')
        lines.append(f'    <Дата>{str(r.get("date",""))[:10]}</Дата>')
        lines.append(f'    <Сумма>{r.get("amount",0)}</Сумма>')
        lines.append(f'    <СчетДт>{r.get("debit_account","")}</СчетДт>')
        lines.append(f'    <СчетКт>{r.get("credit_account","")}</СчетКт>')
        lines.append(f'    <Содержание>{r.get("description","")}</Содержание>')
        lines.append(f'    <Контрагент>{r.get("partner","")}</Контрагент>')
        lines.append('  </ДокументБухгалтерскийУчет>')
    lines.append('  </Документы>')
    lines.append('</ФайлОбменДанными>')
    return "\n".join(lines)


def export_pdf(tenant_id="default", status=None):
    rows = get_journal_drafts(tenant_id, status, limit=500)
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4,
                            rightMargin=20, leftMargin=20,
                            topMargin=30, bottomMargin=20)
    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph(
        f"<b>Bridge Hub — Journal Drafts</b><br/>"
        f"<font size=9>Tenant: {tenant_id} | "
        f"{datetime.now().strftime('%Y-%m-%d %H:%M')} | სულ: {len(rows)}</font>",
        styles["Title"]
    )
    elements.append(title)
    elements.append(Spacer(1, 12))

    table_data = [["ID", "თარიღი", "აღწერა", "თანხა", "Dr", "Cr", "სტატუსი"]]
    for r in rows:
        table_data.append([
            str(r.get("id", "")),
            str(r.get("date", ""))[:10],
            str(r.get("description", ""))[:40],
            str(r.get("amount", "")),
            str(r.get("debit_account", "")),
            str(r.get("credit_account", "")),
            str(r.get("status", "")),
        ])

    t = Table(table_data, colWidths=[35, 60, 160, 55, 35, 35, 75])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F2F2F2"), colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (2, 1), (2, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(t)
    doc.build(elements)
    output.seek(0)
    return output